import os
import gzip
import shutil
import hashlib
import openpyxl
import requests
from pathlib import Path
from io import BytesIO
from collections import defaultdict, OrderedDict
from Bio import SeqIO
from subprocess import Popen, PIPE, CalledProcessError
from datetime import datetime

from django.db import connection
from browser.models import *
from genomebrowser.settings import STATIC_URL

# Register plugins here
from browser.dataimport.plugins.fama import application as fama
from browser.dataimport.plugins.amrfinder import application as amrfinder
from browser.dataimport.plugins.antismash import application as antismash
from browser.dataimport.plugins.phispy import application as phispy
from browser.dataimport.plugins.ecis_screen import application as ecis_screen
""" 
    Various functions for generation and import of gene annotations.
"""

class Annotator(object):

    def __init__(self):
        self.config = {}
        self.read_config()
        self.proteins = {}
        self.hmmsearch_input_file = os.path.join(self.config['cgcms.temp_dir'], 'hmmsearch_input.faa')
        if os.path.exists(self.hmmsearch_input_file):
            os.remove(self.hmmsearch_input_file)
        # Data tables
        self.annotations = []
        self.metadata = []
        
    def read_config(self):
        for item in Config.objects.values('param', 'value'):
            self.config[item['param']] = item['value']

    def export_proteins(self, genome_ids):
        """Populates proteins dictionary and creates FASTA file"""
        if genome_ids is None:
            for item in Protein.objects.all():
                self.proteins[item['protein_hash']] = item
        else:
            target_genes = Gene.objects.filter(genome__id__in = genome_ids).select_related('protein')
            protein_mappings = {}
            for gene in target_genes:
                if gene.protein is not None:
                    self.proteins[gene.protein.protein_hash] = gene.protein
            
        with open(self.hmmsearch_input_file, 'w') as outfile:
            for protein_hash, protein in self.proteins.items():
                outfile.write('>' + protein_hash + '\n')
                outfile.write(protein.sequence + '\n')
        
    def run_hmmsearch(self, outfile, lib_path):
        """
        Runs hmmsearch for FASTA file of proteins.
        """
        # Close MySQL connection before starting external process because it may run for too long resulting in "MySQL server has gone away" error
        connection.close()
        cmd = [self.config['cgcms.hmmsearch_command'],
              '--domtblout', outfile, '-o', '/dev/null',
              '--cut_tc', '--cpu', self.config['cgcms.threads'], '--noali', '--notextw',
              lib_path,
              self.hmmsearch_input_file
              ]
        with Popen(cmd, stdout=PIPE, bufsize=1, universal_newlines=True) as proc:
            for line in proc.stdout:
                print(line, end='')
        if proc.returncode != 0:
            # Suppress false positive no-member error (see https://github.com/PyCQA/pylint/issues/1860)
            # pylint: disable=no-member
            raise CalledProcessError(proc.returncode, proc.args)
        
    def parse_hmmsearch_output(self, hmmsearchfile):
        hits = {}
        with open(hmmsearchfile, 'r') as infile:
            for line in infile:
                if line.startswith('#'):
                    continue
                row = line.rstrip('\n\r').split()
                if (row[0], row[3]) in hits:
                    hits[(row[0], row[3])]['coords'].append(row[17] + '..' + row[18])
                else:
                    hits[(row[0], row[3])] = {'protein_hash':row[0],
                           'hmm_id':row[3],
                           'hmm_acc':row[4],
                           'evalue':row[6],
                           'coords':[row[17] + '..' + row[18]]
                           }
        return hits.values()
        
    def read_hmm_reference(self, ref_path):
        result = {}
        with open(ref_path, 'r') as infile:
            for line in infile:
                if line.startswith('#'):
                    continue
                row = line.rstrip('\n\r').split('\t')
                result[row[0]] = {}
                result[row[0]]['acc'] = row[1]
                result[row[0]]['desc'] = row[-1]
        return result
        
    def add_pfam_domains(self):
        self.proteins = {}
        print('Deleting existing annotations')
        Annotation.objects.filter(source='Pfam database').delete()
        self._create_pfam_domains()
        
    def _create_pfam_domains(self, genome_ids=None):
        """Creates PFAM domain mappings for proteins in the database.
        The genome_ids parameter is a list of genome ids that would be included in hmmsearch search
        """
        batch_size=10000
        # export proteins as FASTA
        print('Exporting proteins')
        self.export_proteins(genome_ids)
        # run hmmscan
        print('Running HMMSEARCH')
        hmmsearch_outfile = os.path.join(self.config['cgcms.temp_dir'], 'hmmsearch_pfam.out.txt')
        if os.path.exists(hmmsearch_outfile):
            os.remove(hmmsearch_outfile)
        self.run_hmmsearch(hmmsearch_outfile, self.config['ref.pfam_hmm_lib'])
        # parse result
        print('Reading HMMSEARCH output')
        hits = self.parse_hmmsearch_output(hmmsearch_outfile)
        # read HMM list
        ref_hmm = self.read_hmm_reference(self.config['ref.pfam_hmm_list'])
        # create Annotations
        print('Creating annotations')
        annotations_written = 0
        for hit in hits:
            protein = self.proteins[hit['protein_hash']]
            for gene in protein.gene_set.all():
                if genome_ids is None or gene.genome.id in genome_ids:
                    self.annotations.append(Annotation(gene_id=gene,
                        source='Pfam database',
                        url='https://www.ebi.ac.uk/interpro/entry/pfam/' + ref_hmm[hit['hmm_id']]['acc'],
                        key='Pfam domain',
                        value=hit['hmm_id'],
                        note=hit['hmm_id'] + ' (' + ref_hmm[hit['hmm_id']]['acc'] + '): ' + ref_hmm[hit['hmm_id']]['desc'] + '. E-value: ' + hit['evalue'] + '. Coordinates: ' + ';'.join(hit['coords'])
                        ))
                    if len(self.annotations) >= batch_size:
                        Annotation.objects.bulk_create(self.annotations, batch_size=batch_size)
                        annotations_written += len(self.annotations)
                        self.annotations = []
        # write Annotations
        print('Writing annotations')
        Annotation.objects.bulk_create(self.annotations, batch_size=batch_size)
        annotations_written += len(self.annotations)
        print(annotations_written, 'annotations created for PFAM domains')
        self.annotations = []
        
    def update_pfam_domains(self, genome_ids=None):
        self.proteins = {}
        print('Delete existing PFAM annotations')
        if genome_ids is None:
            Annotation.objects.filter(source='Pfam database').delete()
        else:
            Annotation.objects.filter(source='Pfam database', gene_id__genome__id__in=genome_ids).delete()
        self._create_pfam_domains(genome_ids)

    def update_tigrfam_domains(self, genome_ids=None):
        self.proteins = {}
        print('Delete existing TIGRFAM annotations')
        if genome_ids is None:
            Annotation.objects.filter(source='TIGRFAM database').delete()
        else:
            Annotation.objects.filter(source='TIGRFAM database', gene_id__genome__id__in=genome_ids).delete()
        self._create_tigrfam_domains(genome_ids)

    def add_tigrfam_domains(self):
        self.proteins = {}
        print('Deleting existing annotations')
        Annotation.objects.filter(source='TIGRFAM database').delete()
        self._create_tigrfam_domains()
    
    def _create_tigrfam_domains(self, genome_ids=None):
        batch_size=10000
        print('Exporting proteins')
        self.export_proteins(genome_ids)
        # run hmmscan
        print('Running HMMSEARCH')
        hmmsearch_outfile = os.path.join(self.config['cgcms.temp_dir'], 'hmmsearch_tigrfam.out.txt')
        if os.path.exists(hmmsearch_outfile):
            os.remove(hmmsearch_outfile)
        self.run_hmmsearch(hmmsearch_outfile, self.config['ref.tigrfam_hmm_lib'])
        # parse result
        print('Reading HMMSEARCH output')
        hits = self.parse_hmmsearch_output(hmmsearch_outfile)
        # read HMM list
        ref_hmm = self.read_hmm_reference(self.config['ref.tigrfam_hmm_list'])
        # create Annotations
        print('Creating annotations')
        annotations_written = 0
        for hit in hits:
            protein = self.proteins[hit['protein_hash']]
            for gene in protein.gene_set.all():
                if genome_ids is None or gene.genome.id in genome_ids:
                    self.annotations.append(Annotation(gene_id=gene,
                        source='TIGRFAM database',
                        url='http://tigrfams.jcvi.org/cgi-bin/HmmReportPage.cgi?acc=' + hit['hmm_id'],
                        key='TIGRFAM family',
                        value=hit['hmm_id'],
                        note=hit['hmm_id'] + ': ' + ref_hmm[hit['hmm_id']]['desc'] + '. E-value: ' + hit['evalue'] + '. Coordinates: ' + ';'.join(hit['coords'])
                        ))
                    if len(self.annotations) >= batch_size:
                        Annotation.objects.bulk_create(self.annotations, batch_size=batch_size)
                        annotations_written += len(self.annotations)
                        self.annotations = []
        # write Annotations
        print('Writing annotations')
        Annotation.objects.bulk_create(self.annotations, batch_size=10000)
        annotations_written += len(self.annotations)
        print(annotations_written, 'annotations created for TIGRFAM families')
        self.annotations = []
    
    def make_fitbrowser_annotations(self, strain, org_name, protein_path):
        result = []
        for seq_record in SeqIO.parse(protein_path, "fasta"):
            protein_sequence = seq_record.seq
            # Some protein sequences were translated incorrectly
            if protein_sequence.startswith('V') or protein_sequence.startswith('L'):
                protein_sequence = 'M' + protein_sequence[1:]
            protein_hash = hashlib.md5(protein_sequence.encode('utf-8')).hexdigest()
            #print(protein_hash, seq_record.id)
            #print(seq_record.seq)
            if protein_hash in self.proteins:
                #print(protein_hash + 'protein found!')
                seq_id = seq_record.id.split(' ')[0]
                org_id, locus_id = seq_id.split(':')
                #print(org_id, locus_id)
                for gene in self.proteins[protein_hash].gene_set.all():
                    if gene.genome.strain == strain:
                        #print('Adding annotation to gene', gene.locus_tag)
                        result.append(Annotation(gene_id=gene,
                            source='Fitness Browser',
                            url='http://fit.genomics.lbl.gov/cgi-bin/singleFit.cgi?orgId=' + org_id + '&locusId=' + locus_id,
                            key='Fitness Browser gene',
                            value=locus_id,
                            note='Gene ' + locus_id + ' from ' + org_name
                            ))
        print(org_name, 'genes found:', len(result))
        return result
        
    def add_fitbrowser_links(self):
        """ This function adds gene annotations with links to Fitness Browser (http://fit.genomics.lbl.gov/)"""
        self.annotations = []
        # Delete existing annotations
        print('Deleting existing annotations')
        Annotation.objects.filter(source='Fitness Browser').delete()
        print('Making protein list')
        for protein in Protein.objects.all():
            self.proteins[protein.protein_hash] = protein
        strains = {item.strain_id:item for item in Strain.objects.all()}
        print('Reading organism list')
        with open(self.config['ref.fitbrowser_orgs_file'], 'r') as infile:
            for line in infile:
                row = line.rstrip('\n\r').split('\t')
                if row[0] in strains:
                    print(row[0], 'found in database')
                    self.annotations += self.make_fitbrowser_annotations(strains[row[0]], row[1], row[3])
                    # write Annotations
                    print('Writing annotations')
                    Annotation.objects.bulk_create(self.annotations, batch_size=10000)
                    print(len(self.annotations), 'annotations created for', row[0])
                    self.annotations = []
        self.annotations = []

    def add_custom_annotations(self, tsv_file):
        """ This function adds gene annotations from tab-separated file. 
        If such anotation already exists, the existing copy will be preserved (annotation is considered identical if, for the same gene, it has same source, key and value)"""
        print('Reading annotations from file')
        lines = []
        with open(tsv_file, 'r') as infile:
            for line in infile:
                if line.startswith('#'):
                    continue
                lines.append(line)
        self.import_annotations(lines)
        
    def import_annotations(self, lines):
        """ This function adds gene annotations from a list of lines. 
        If an anotation already exists, the existing copy will be preserved (annotation is considered identical if, for the same gene, it has same source, key and value)"""
        batch_size=10000
        annotations_written = 0
        for line in lines:
            if line.startswith('#'):
                continue
            locus_tag, genome_name, source, url, key, value, note = line.rstrip('\n\r').split('\t')
            existing_annotations = Annotation.objects.filter(gene_id__locus_tag = locus_tag,
                gene_id__genome__name = genome_name,
                source=source,
                key=key,
                value=value)
            if existing_annotations:
                print('Annotation exists', locus_tag, genome_name, source, key, value)
            else:
                try:
                    gene = Gene.objects.get(locus_tag=locus_tag, genome__name = genome_name)
                    self.annotations.append(Annotation(gene_id=gene,
                        source=source,
                        url=url,
                        key=key,
                        value=value,
                        note=note
                        ))
                    print('Annotation added', locus_tag, genome_name, source, key, value)
                except Gene.DoesNotExist:
                    print('Gene not found', locus_tag, genome_name)
            if len(self.annotations) >= batch_size:
                Annotation.objects.bulk_create(self.annotations, batch_size=batch_size)
                annotations_written += len(self.annotations)
                self.annotations = []
        # write Annotations
        if self.annotations:
            Annotation.objects.bulk_create(self.annotations, batch_size=10000)
        annotations_written += len(self.annotations)
        print(annotations_written, 'annotations written')
        self.annotations = []

    def add_regulons(self, lines):
        """ This function adds sites and regulons from a list of tab-separated lines. 
        If such a site already exists, the existing copy will be preserved (site is considered identical if it has the same genome, contig. start, end and strand)
        """
        print('Reading regulons from file')
        regulon_data = autovivify(2, list)
        
        for line in lines:
            regulon_name, genome_name, reg_gene_ids, target_gene_id, contig, start, end, strand, sequence = line.rstrip('\n\r').split('\t')
            # check if genes exist
            try:
                for regulator_id in reg_gene_ids.split(','):
                    gene = Gene.objects.get(locus_tag=regulator_id, genome__name=genome_name)
                gene = Gene.objects.get(locus_tag=target_gene_id, genome__name=genome_name)
            except Gene.DoesNotExist:
                continue
            regulon_data[genome_name][regulon_name].append((reg_gene_ids, target_gene_id, contig, start, end, strand, sequence))
        print(regulon_data)
        
        for genome_name in regulon_data:
            genome = Genome.objects.get(name=genome_name)
            # Create regulons
            existing_regulons = {item.name:item for item in Regulon.objects.filter(genome__name = genome_name)}
            for regulon_name in regulon_data[genome_name]:
                print(regulon_name, genome_name)
                if regulon_name not in existing_regulons:
                    regulator_ids = regulon_data[genome_name][regulon_name][0][0].split(',')
                    regulon = Regulon(name=regulon_name, genome=genome)
                    regulon.save()
                    existing_regulons[regulon_name] = regulon
                    print(str(regulon))
                    for regulator_id in regulator_ids:
                        regulon.regulators.add(Gene.objects.get(locus_tag=regulator_id, genome__name=genome_name))
                else:
                    regulon = existing_regulons[regulon_name]
                    existing_regulators = [item.locus_tag for item in regulon.regulators.all()]
                    for regulon_ind in regulon_data[genome_name][regulon_name]:
                        regulator_ids = regulon_ind[0].split(',')
                        for regulator_id in regulator_ids:
                            print (regulator_id)
                            if regulator_id not in existing_regulators:
                                regulon.regulators.add(Gene.objects.get(locus_tag=regulator_id, genome__name=genome_name))
                                existing_regulators.append(regulator_id)
            # Create sites
            existing_regulons = {item.name:item for item in Regulon.objects.filter(genome__name = genome_name)}
            existing_sites = {item.name:item for item in Site.objects.filter(genome__name = genome_name)}
            for regulon_name in regulon_data[genome_name]:
                for site_data in regulon_data[genome_name][regulon_name]:
                    site_name = regulon_name + '_site_at_' + site_data[1]
                    if site_name in existing_sites:
                        site_index = 2
                        while True:
                            site_name = regulon_name + '_site_' + str(site_index) + '_at_' + site_data[1]
                            if site_name not in existing_sites:
                                break
                            site_index += 1
                    target_gene = Gene.objects.get(locus_tag=site_data[1], genome=genome)
                    target_operon = target_gene.operon
                    same_sites = Site.objects.filter(genome__name = genome_name, contig__contig_id = site_data[2], start = int(site_data[3]), end = int(site_data[4]), strand = int(site_data[5]))
                    if not same_sites:
                        print(site_data)
                        site = Site(name = site_name, type = 'TFBS', start = int(site_data[3]), end = int(site_data[4]), strand = int(site_data[5]), contig = Contig.objects.get(genome__name = genome_name, contig_id = site_data[2]), genome = genome, sequence = site_data[6], regulon = existing_regulons[regulon_name])
                        site.save()
                        existing_sites[site_name] = site
                        if target_operon:
                            site.operons.add(target_operon)
                        else:
                            site.genes.add(target_gene)

    def add_strain_metadata(self, tsv_file):
        """ This function adds strain metadata from tab-separated file
        N.B.: This function no longer associated with add_strain_metadata management command.
        """
        print('Reading metadata from file')
        self.metadata = []
        strains = {item.strain_id:item for item in Strain.objects.all()}
        with open(tsv_file, 'r') as infile:
            for line in infile:
                if line.startswith('#'):
                    continue
                strain, source, url, key, value = line.rstrip('\n\r').split('\t')
                if strain in strains:
                    self.metadata.append(Strain_metadata(strain=strains[strain],
                        source=source,
                        url=url,
                        key=key,
                        value=value
                        ))
                    print('Metadata added', strain, source, key, value)
        # write metadata
        print('Writing metadata')
        Strain_metadata.objects.bulk_create(self.metadata, batch_size=10000)
        print(len(self.metadata), 'metadata entries written')

    def update_strain_metadata(self, xlsx_path=None, xlsx_file=None):
        """ This function adds strain metadata from Excel file and from isolates.genomics.lbl.gov API"""
        print('Reading metadata from file', xlsx_file)
        metadata_imported = defaultdict(dict)
        if xlsx_path is not None:
            xlsx_path = Path(xlsx_path)
            wb_obj = openpyxl.load_workbook(xlsx_path)
        else:
            wb_obj = openpyxl.load_workbook(filename=BytesIO(xlsx_file.read()))
        sheet = wb_obj.active
        xlsx_header = []

        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i == 0:
                xlsx_header = row[1:]
            else:
                strain_id = row[0]
                if strain_id == '' or strain_id is None:
                    continue
                for j, cell in enumerate(row[1:]):
                    if cell != '' and cell != 'None' and cell is not None:
                        metadata_imported[strain_id][xlsx_header[j]] = ('User-defined data', 'javascript:alert(\'No external link.\');', str(cell))
        
        # Download from isolates.genomics.lbl.gov 
        print('Downloading metadata from isolates.genomics.lbl.gov')
        url = 'http://isolates.genomics.lbl.gov/api/v1/isolates/id/'
        external_url = 'http://isolates.genomics.lbl.gov/isolates/id/'
        fields = [('condition', 'Isolation conditions/description (including temperature)'),
                  ('order', 'Phylogenetic Order'),
                  ('closest_relative', 'Closest relative in NCBI: 16S rRNA Gene Database'),
                  ('similarity', 'Similarity (%)'),
                  ('date_sampled', 'Date sampled'),
                  ('sample_id', 'Well/Sample ID'),
                  ('lab', 'Lab isolated/Contact'),
                  ('campaign', 'Campaign or Set'),
                  ('rrna', 'rrna')
                  ]
        error_threshold = 100
        isolate_max_id = 100000
        errors = 0
        for isolate_id in range(isolate_max_id):
            isolate_url = url + str(isolate_id)
            r = requests.get(url=isolate_url)
            # extracting data in json format 
            data = r.json()
            if 'id' not in data:
                print(isolate_url, 'returned no data')
                errors += 1
            if errors >= error_threshold:
                print('Download stopped after', error_threshold, 'errors. Last url is', isolate_url)
                break
            try:
                print(data['id'], data['isolate_id'])
                strain_id = data['isolate_id']
                for field in fields:
                    if field[0] in data:
                        if data[field[0]] is not None and data[field[0]] != '':
                            metadata_imported[strain_id][field[1]] = ('ENIGMA Isolate Browser', external_url + str(isolate_id), str(data[field[0]]))
            except KeyError:
                continue
        # Export metadata
        metadata_file = self.config['strains.metadata_file']
        with open(metadata_file, 'w') as outfile:
            for strain_id in sorted(metadata_imported.keys()):
                for key in sorted(metadata_imported[strain_id].keys()):
                    outfile.write('\t'.join([strain_id,
                                             metadata_imported[strain_id][key][0],
                                             metadata_imported[strain_id][key][1],
                                             key,
                                             metadata_imported[strain_id][key][2]]) + '\n')
        # Check existing entries
        print('Updating exisiting metadata')
        strains = {item.strain_id:item for item in Strain.objects.all()}
        for strain_id in strains:
            if strain_id not in metadata_imported:
                continue
            strain_metadata = {item.key: item for item in Strain_metadata.objects.filter(strain__strain_id=strain_id)}
            for key in metadata_imported[strain_id]:
                if key in strain_metadata:
                    if strain_metadata[key].value != metadata_imported[strain_id][key][2]:
                        strain_metadata[key].value = metadata_imported[strain_id][key][2]
                        strain_metadata[key].save()
                    if strain_metadata[key].source != metadata_imported[strain_id][key][0]:
                        strain_metadata[key].source = metadata_imported[strain_id][key][0]
                        strain_metadata[key].save()
                    if strain_metadata[key].url != metadata_imported[strain_id][key][1]:
                        strain_metadata[key].url = metadata_imported[strain_id][key][1]
                        strain_metadata[key].save()
                else:
                    new_metadata = Strain_metadata(strain=strains[strain_id],
                        source=metadata_imported[strain_id][key][0],
                        url=metadata_imported[strain_id][key][1],
                        key=key,
                        value=metadata_imported[strain_id][key][2]
                        )
                    new_metadata.save()

    def add_sample_metadata(self, lines):
        """ This function adds sample metadata from tab-separated file"""
        self.metadata = []
        samples = {item.sample_id:item for item in Sample.objects.all()}
        for line in lines:
            if line.startswith('#'):
                continue
            sample, source, url, key, value = line.rstrip('\n\r').split('\t')
            if sample in samples:
                self.metadata.append(Sample_metadata(sample=samples[sample],
                    source=source,
                    url=url,
                    key=key,
                    value=value
                    ))
                print('Metadata added', sample, source, key, value)
        # write metadata
        print('Writing metadata')
        Sample_metadata.objects.bulk_create(self.metadata, batch_size=10000)
        print(len(self.metadata), 'metadata entries written')

    def update_genome_descriptions(self, tsv_file):
        """ This function reads genome descriptions from tab-separated file and updates Genome objects"""
        print('Reading metadata from file')
        genomes = {item.name:item for item in Genome.objects.all()}
        with open(tsv_file, 'r') as infile:
            for line in infile:
                if line.startswith('#'):
                    continue
                genome_name, description = line.rstrip('\n\r').split('\t')
                if genome_name in genomes:
                    genomes[genome_name].description = description
                    genomes[genome_name].save()
                    print('Genome description updated for', genome_name)

    def update_sample_descriptions(self, lines):
        """ This function reads metagenomic sample descriptions from tab-separated file and updates Sample objects"""
        samples = {item.sample_id:item for item in Sample.objects.all()}
        for line in lines:
            if line.startswith('#'):
                continue
            sample_name, full_name, description = line.rstrip('\n\r').split('\t')
            if sample_name in samples:
                samples[sample_name].description = description
                samples[sample_name].full_name = full_name
                samples[sample_name].save()
                print('Sample description updated for', sample_name)

    def run_external_tools(self, genomes):
        plugins_available = set()
        for param in self.config:
            if param.startswith('plugins.') and self.config[param] != '':
                plugin = param.split('.')[1]
                plugins_available.add(plugin)
                
        print('Available plugins:', ','.join(list(plugins_available)))

        # Run Fama
        if 'fama' in plugins_available:
            fama_annotations = fama(self, genomes)
            print(fama_annotations, 'ready for upload')
            self.add_custom_annotations(fama_annotations)

        # Run amrfinder
        if 'amrfinder' in plugins_available:
            amrfinder_annotations = amrfinder(self, genomes)
            print(amrfinder_annotations, 'ready for upload')
            self.add_custom_annotations(amrfinder_annotations)

        # Run antiSMASH
        if 'antismash' in plugins_available:
            antismash_annotations = antismash(self, genomes)
            print(antismash_annotations, 'ready for upload')
            self.add_custom_annotations(antismash_annotations)
        
        # Run PhiSpy
        if 'phispy' in plugins_available:
            phispy_annotations = phispy(self, genomes)
            print(phispy_annotations, 'ready for upload')
            self.add_custom_annotations(phispy_annotations)

        # Run eCIS-screen
        if 'ecis_screen' in plugins_available:
            ecis_screen_annotations = ecis_screen(self, genomes)
            print(ecis_screen_annotations, 'ready for upload')
            self.add_custom_annotations(ecis_screen_annotations)
    
                
def autovivify(levels=1, final=dict):
    return (defaultdict(final) if levels < 2 else
            defaultdict(lambda: autovivify(levels - 1, final)))
