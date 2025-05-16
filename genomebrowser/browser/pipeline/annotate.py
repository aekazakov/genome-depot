import os
import re
import openpyxl
import importlib
import pkgutil
import logging
import traceback
from pathlib import Path
from io import BytesIO
from collections import defaultdict
from browser.models import Annotation
from browser.models import Config
from browser.models import Contig
from browser.models import Gene
from browser.models import Genome
from browser.models import Regulon
from browser.models import Sample
from browser.models import Sample_metadata
from browser.models import Site
from browser.models import Strain
from browser.models import Strain_metadata
from genomebrowser.settings import BASE_DIR
from browser.pipeline.util import autovivify
import browser.pipeline.plugins

""" 
    Various functions for generation and import of gene annotations.
"""

logger = logging.getLogger("GenomeDepot")

class Annotator(object):

    def __init__(self):
        self.config = {}
        self._read_config()
        self.proteins = {}
        # Data tables
        self.annotations = []
        self.metadata = []
        self.plugins = {name: importlib.import_module(name)
                        for finder, name, ispkg
                        in pkgutil.iter_modules(browser.pipeline.plugins.__path__,
                                                browser.pipeline.plugins.__name__ + "."
                                                )
                        if name.startswith('browser.pipeline.plugins.genomedepot_')
                        }
        
        
    def _read_config(self):
        """
            Reads all Config entries from database into the self.config dictionary
        """
        for item in Config.objects.values('param', 'value'):
            self.config[item['param']] = item['value']
    
    def add_custom_annotations(self, tsv_file):
        """ This function adds gene annotations from tab-separated file. 
        If such anotation already exists, the existing copy will be 
        preserved (annotation is considered identical if, for 
        the same gene, it has same source, key and value)
        """
        logger.info('Reading annotations from file')
        lines = []
        with open(tsv_file, 'r') as infile:
            for line in infile:
                if line.startswith('#'):
                    continue
                lines.append(line)
        self.import_annotations(lines)
        
    def import_annotations(self, lines):
        """ This function adds gene annotations from a list of lines. 
        If an anotation already exists, the existing copy will be preserved 
        (annotation is considered identical if, for the same gene,
        it has same source, key and value)
        """
        batch_size=10000
        annotations_written = 0
        for line in lines:
            if line.startswith('#'):
                continue
            locus_tag, genome_name, source, url, key, value, note = \
                line.rstrip('\n\r').split('\t')
            if len(value) > 50:
                value = value[:47] + '...'
            existing_annotations = \
                Annotation.objects.filter(
                                          gene_id__locus_tag = locus_tag,
                                          gene_id__genome__name = genome_name,
                                          source=source,
                                          key=key,
                                          value=value
                                          )
            if existing_annotations:
                logger.info(' '.join(['Annotation exists',
                                      locus_tag,
                                      genome_name,
                                      source,
                                      key,
                                      value
                                      ]))
            else:
                try:
                    gene = Gene.objects.get(locus_tag=locus_tag,
                                            genome__name = genome_name
                                            )
                    self.annotations.append(Annotation(gene_id=gene,
                        source=source,
                        url=url,
                        key=key,
                        value=value,
                        note=note
                        ))
                    logger.info(' '.join(['Annotation added',
                                          locus_tag,
                                          genome_name,
                                          source,
                                          key,
                                          value
                                          ]))
                except Gene.DoesNotExist:
                    logger.warning('Gene %s not found in %s', locus_tag, genome_name)
            if len(self.annotations) >= batch_size:
                try:
                    Annotation.objects.bulk_create(self.annotations,
                                                   batch_size=batch_size
                                                   )
                except Exception:
                    logger.info(
                        'An error occurred while saving annotations to the database'
                    )
                    self.annotations = []
                    raise
                annotations_written += len(self.annotations)
                self.annotations = []
        # write Annotations
        if self.annotations:
            try:
                Annotation.objects.bulk_create(self.annotations, batch_size=10000)
            except Exception:
                logger.info(
                    'An error occurred while saving annotations to the database'
                )
                self.annotations = []
                raise
        annotations_written += len(self.annotations)
        logger.info('%d annotations written', annotations_written)
        self.annotations = []

    def add_regulons(self, lines):
        """ 
        This function adds sites and regulons from a list of 
        tab-separated lines. 
        If such a site already exists, the existing copy will be preserved 
        (site is considered identical if it has the same genome, 
        contig, start, end and strand)
        """
        logger.info('Reading regulons from file')
        regulon_data = autovivify(2, list)
        
        for line in lines:
            regulon_name, genome_name, reg_gene_ids, target_gene_id, contig, \
            start, end, strand, sequence = line.rstrip('\n\r').split('\t')
            regulon_name = self.sanitize_input(regulon_name)
            # check if genes exist
            skip_line = False
            for regulator_id in reg_gene_ids.split(','):
                if not Gene.objects.filter(locus_tag=regulator_id,
                                           genome__name=genome_name
                                           ).exists():
                    skip_line = True
            if not Gene.objects.filter(locus_tag=target_gene_id,
                                       genome__name=genome_name
                                       ).exists():
                skip_line = True
            if skip_line:
                continue
            regulon_data[genome_name][regulon_name].append((reg_gene_ids,
                                                            target_gene_id,
                                                            contig, 
                                                            start, 
                                                            end, 
                                                            strand, 
                                                            sequence
                                                            ))
        for genome_name in regulon_data:
            genome = Genome.objects.get(name=genome_name)
            # Create regulons
            existing_regulons = {item.name:item for item
                                 in Regulon.objects.filter(genome__name = genome_name)
                                 }
            for regulon_name in regulon_data[genome_name]:
                logger.info(regulon_name + ' ' + genome_name)
                if regulon_name not in existing_regulons:
                    regulator_ids = \
                        regulon_data[genome_name][regulon_name][0][0].split(',')
                    regulon = Regulon(name=regulon_name, genome=genome)
                    regulon.save()
                    existing_regulons[regulon_name] = regulon
                    logger.info(str(regulon))
                    for regulator_id in regulator_ids:
                        regulon.regulators.add(
                            Gene.objects.get(locus_tag=regulator_id,
                                             genome__name=genome_name)
                            )
                else:
                    regulon = existing_regulons[regulon_name]
                    existing_regulators = [item.locus_tag for item 
                                           in regulon.regulators.all()]
                    for regulon_ind in regulon_data[genome_name][regulon_name]:
                        regulator_ids = regulon_ind[0].split(',')
                        for regulator_id in regulator_ids:
                            logger.info(regulator_id)
                            if regulator_id not in existing_regulators:
                                regulon.regulators.add(
                                    Gene.objects.get(locus_tag=regulator_id,
                                                     genome__name=genome_name
                                                     )
                                    )
                                existing_regulators.append(regulator_id)
            # Create sites
            existing_regulons = {item.name:item for item
                                 in Regulon.objects.filter(
                                    genome__name = genome_name
                                    )
                                 }
            existing_sites = {item.name:item for item
                              in Site.objects.filter(genome__name = genome_name)
                              }
            for regulon_name in regulon_data[genome_name]:
                for site_data in regulon_data[genome_name][regulon_name]:
                    site_name = regulon_name + '_site_at_' + site_data[1]
                    if site_name in existing_sites:
                        site_index = 2
                        while True:
                            site_name = regulon_name + '_site_' + str(site_index) + \
                                        '_at_' + site_data[1]
                            if site_name not in existing_sites:
                                break
                            site_index += 1
                    target_gene = Gene.objects.get(locus_tag=site_data[1],
                                                   genome=genome
                                                   )
                    target_operon = target_gene.operon
                    same_sites = Site.objects.filter(genome__name = genome_name,
                                                     contig__contig_id = site_data[2],
                                                     start = int(site_data[3]),
                                                     end = int(site_data[4]),
                                                     strand = int(site_data[5])
                                                     )
                    if not same_sites:
                        logger.info(str(site_data))
                        try:
                            target_contig = Contig.objects.get(
                                        genome__name = genome_name,
                                        contig_id = site_data[2]
                                        )
                        except Contig.DoesNotExist:
                            logging.error(site_data[2] + 
                                ' contig not found in ' + genome_name
                                )
                            raise
                        site = Site(name = site_name,
                                    type = 'TFBS',
                                    start = int(site_data[3]),
                                    end = int(site_data[4]),
                                    strand = int(site_data[5]),
                                    contig = target_contig,
                                    genome = genome,
                                    sequence = site_data[6],
                                    regulon = existing_regulons[regulon_name]
                                    )
                        site.save()
                        existing_sites[site_name] = site
                        if target_operon:
                            site.operons.add(target_operon)
                        else:
                            site.genes.add(target_gene)

    def add_strain_metadata(self, tsv_file):
        """ This function adds strain metadata from tab-separated file
            N.B.: This function is not associated with add_strain_metadata
            management command.
        """
        if tsv_file is None or not os.path.exists(tsv_file):
            logger.error('Input file does not exists')
            return
        logger.info('Reading metadata from file')
        self.metadata = []
        strains = {item.strain_id:item for item in Strain.objects.all()}
        with open(tsv_file, 'r') as infile:
            for line in infile:
                if line.startswith('#'):
                    continue
                strain, source, url, key, value = line.rstrip('\n\r').split('\t')
                if strain in strains:
                    self.metadata.append(Strain_metadata(strain=strains[strain],
                        source=source,
                        url=url,
                        key=key,
                        value=value
                        ))
                    logger.info('Metadata added %s %s %s %s',
                                strain, source, key, value
                                )
        # write metadata
        logger.info('Writing strain metadata')
        try:
            Strain_metadata.objects.bulk_create(self.metadata, batch_size=10000)
            logger.info('%s metadata entries written', len(self.metadata))
            self.metadata = []
        except Exception:
            logger.error('An error occurred while saving metadata to the database')
            self.metadata = []
            raise

    def update_strain_metadata(self, xlsx_path=None, xlsx_file=None):
        """ 
            This function adds strain metadata from an Excel file.
            If the enigma module is configured, also imports 
            from isolates.genomics.lbl.gov API.
        """
        if xlsx_path is None and xlsx_file is None:
            logger.error('No input data')
            return
        metadata_imported = defaultdict(dict)
        if xlsx_path is not None:
            logger.info('Reading metadata from file %s', xlsx_path)
            xlsx_path = Path(xlsx_path)
            wb_obj = openpyxl.load_workbook(xlsx_path)
        else:
            logger.info('Reading metadata from file %s', xlsx_file)
            wb_obj = openpyxl.load_workbook(filename=BytesIO(xlsx_file.read()))
        sheet = wb_obj.active
        xlsx_header = []

        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i == 0:
                xlsx_header = row[1:]
            else:
                strain_id = row[0]
                if strain_id == '' or strain_id is None:
                    continue
                for j, cell in enumerate(row[1:]):
                    if cell != '' and cell != 'None' and cell is not None:
                        metadata_imported[strain_id][xlsx_header[j]] = \
                        ('User-defined data',
                        'javascript:alert(\'No external link.\');',
                        str(cell)
                        )
        
        if os.path.exists(os.path.join(BASE_DIR, 'enigma', 'enigma.py')):
            from enigma.enigma import download_enigma_strain_metadata
            isolate_browser_metadata = download_enigma_strain_metadata()
            metadata_imported.update(isolate_browser_metadata)

        # Check existing entries
        logger.info('Updating exisiting metadata')
        strains = {item.strain_id:item for item in Strain.objects.all()}
        for strain_id in strains:
            if strain_id not in metadata_imported:
                continue
            strain_metadata = {item.key: item for item 
                               in Strain_metadata.objects.filter(
                                    strain__strain_id=strain_id
                                    )
                               }
            for key in metadata_imported[strain_id]:
                if key in strain_metadata:
                    if strain_metadata[key].value != \
                    metadata_imported[strain_id][key][2]:
                        strain_metadata[key].value = \
                            metadata_imported[strain_id][key][2]
                        strain_metadata[key].save()
                    if strain_metadata[key].source != \
                    metadata_imported[strain_id][key][0]:
                        strain_metadata[key].source = \
                            metadata_imported[strain_id][key][0]
                        strain_metadata[key].save()
                    if strain_metadata[key].url != \
                    metadata_imported[strain_id][key][1]:
                        strain_metadata[key].url = metadata_imported[strain_id][key][1]
                        strain_metadata[key].save()
                else:
                    new_metadata = Strain_metadata(strain=strains[strain_id],
                        source=metadata_imported[strain_id][key][0],
                        url=metadata_imported[strain_id][key][1],
                        key=key,
                        value=metadata_imported[strain_id][key][2]
                        )
                    new_metadata.save()

    def add_sample_metadata(self, lines):
        """ This function adds sample metadata from tab-separated file"""
        self.metadata = []
        samples = {item.sample_id:item for item in Sample.objects.all()}
        for line in lines:
            if line.startswith('#'):
                continue
            sample, source, url, key, value = line.rstrip('\n\r').split('\t')
            if sample in samples:
                self.metadata.append(Sample_metadata(sample=samples[sample],
                    source=source,
                    url=url,
                    key=key,
                    value=value
                    ))
                logger.info('Metadata added %s %s %s %s', sample, source, key, value)
        # write metadata
        logger.info('Writing sample metadata')
        try:
            Sample_metadata.objects.bulk_create(self.metadata, batch_size=10000)
            logger.info('%d metadata entries written', len(self.metadata))
            self.metadata = []
        except Exception:
            logger.error('An error occurred while saving metadata to the database')
            self.metadata = []
            raise
            

    def update_genome_descriptions(self, tsv_file):
        """ 
            Reads genome descriptions from a tab-separated file
            and updates Genome objects
        """
        logger.info('Reading metadata from file')
        genomes = {item.name:item for item in Genome.objects.all()}
        with open(tsv_file, 'r') as infile:
            for line in infile:
                if line.startswith('#'):
                    continue
                genome_name, description = line.rstrip('\n\r').split('\t')
                if genome_name in genomes:
                    genomes[genome_name].description = description
                    genomes[genome_name].save()
                    logger.info('Genome description updated for %s', genome_name)

    def update_sample_descriptions(self, lines):
        """ 
            Reads metagenomic sample descriptions from lines of tab-separated 
            file and updates Sample objects
        """
        samples = {item.sample_id:item for item in Sample.objects.all()}
        for line in lines:
            if line.startswith('#'):
                continue
            sample_name, full_name, description = line.rstrip('\n\r').split('\t')
            if sample_name in samples:
                samples[sample_name].description = description
                samples[sample_name].full_name = full_name
                samples[sample_name].save()
                logger.info('Sample description updated for %s', sample_name)

    def run_external_tools(self, genomes, plugin_name=''):
        """
            Runs either one or all external annotation tools for a number of genomes
        """
        ret_val = 'not started'
                
        plugins_available = self.plugins
        logger.info('Starting %s', plugin_name)
        genome_names = list(genomes.keys())
        plugin_module_name = 'browser.pipeline.plugins.genomedepot_' + plugin_name

        if plugin_module_name in plugins_available:
            plugin = plugins_available[plugin_module_name]
            plugin_output = plugin.application(self, genomes)
            logger.info('%s finished', plugin_name)
            display_name = self.config['plugins.' + plugin_name + '.display_name']
            Annotation.objects.filter(source=display_name,
                                      gene_id__genome__name__in=genome_names
                                      ).delete()
            self.add_custom_annotations(plugin_output)
            logger.info('%s output imported', plugin_name)
            ret_val = 'OK'
        else:
            logger.warning('%s module not found. Skipping plugin', plugin_name)
            logger.info('Available modules: %s', str(plugins_available))
            ret_val = 'skipped'
        return ret_val

    def run_annotation_pipeline(self, genomes):
        """
            Runs either one or all external annotation tools for a number of genomes
        """
        ret = []
        plugins_enabled = set()
        for param in self.config:
            if param.startswith('plugins.') and param.endswith('.enabled') and \
            self.config[param] in ('1', 'yes', 'Yes', 'y', 'Y'):
                tool = param.split('.')[1]
                plugins_enabled.add(tool)
                
        plugin_count = 0
        for plugin_name in plugins_enabled:
            plugin_count += 1
            try:
                ret.append(
                    str(plugin_count) + ' of ' + str(len(plugins_enabled)) +
                    ': ' + plugin_name + ' ' +
                    self.run_external_tools(genomes, plugin_name)
                )
            except Exception:
                ret.append(str(plugin_count) + ' of ' + str(len(plugins_enabled)) +
                    ': ' + plugin_name + ' finished with error'
                    )
                ret.append(traceback.format_exc())
                continue

        return '\n'.join(ret)

    def sanitize_input(self, input_text):
        """
            Replaces non-ASCII and non-word characters with 
            underscores in input_text
        """
        input_text = ''.join([i if ord(i) < 128 else ' ' for i in input_text])
        input_text = re.sub(r"[^^a-zA-Z0-9\.\-]", ' ', input_text)
        input_text = "_".join( input_text.split() )
        return input_text
